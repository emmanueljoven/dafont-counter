import time
import re
import csv
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


BASE_URL = "https://www.dafont.com"
THEME_URL = "https://www.dafont.com/es/theme.php"

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

OUTPUT_CSV = DATA_DIR / "dafont_fuentes.csv"
OUTPUT_XLSX = DATA_DIR / "dafont_dataset.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}


def get_categories():
    return [
        ("Fantasía", "Animación", 101),
        ("Fantasía", "Tebeo, Cómic", 102),
        ("Fantasía", "Groovy", 103),
        ("Fantasía", "Old School", 104),
        ("Fantasía", "Rizos", 105),
        ("Fantasía", "Oeste", 106),
        ("Fantasía", "Desgastado", 107),
        ("Fantasía", "Distorsionado", 108),
        ("Fantasía", "Destrozado", 109),
        ("Fantasía", "Terror", 110),
        ("Fantasía", "Fuego, Hielo", 111),
        ("Fantasía", "Decoración", 112),
        ("Fantasía", "Máquina de escribir", 113),
        ("Fantasía", "Stencil, Army", 114),
        ("Fantasía", "Retro", 115),
        ("Fantasía", "Iniciales", 116),
        ("Fantasía", "Cuadrícula", 117),
        ("Fantasía", "Varios", 118),

        ("Aspecto extranjero", "Chino, Japonés", 201),
        ("Aspecto extranjero", "Árabe", 202),
        ("Aspecto extranjero", "Mexicano", 203),
        ("Aspecto extranjero", "Romano, Griego", 204),
        ("Aspecto extranjero", "Ruso", 205),
        ("Aspecto extranjero", "Varios", 206),

        ("Tecno", "Cuadrado", 301),
        ("Tecno", "LCD", 302),
        ("Tecno", "Ciencia ficción", 303),
        ("Tecno", "Varios", 304),

        ("Gótico", "Medieval", 401),
        ("Gótico", "Moderno", 402),
        ("Gótico", "Celta", 403),
        ("Gótico", "Iniciales", 404),
        ("Gótico", "Varios", 405),

        ("Básico", "Sans serif", 501),
        ("Básico", "Serif", 502),
        ("Básico", "Ancho fijo", 503),
        ("Básico", "Varios", 504),

        ("Script", "Caligrafía", 601),
        ("Script", "Escolar", 602),
        ("Script", "Manuscrito", 603),
        ("Script", "Brocha, Pincel", 604),
        ("Script", "Garabato", 605),
        ("Script", "Graffiti", 606),
        ("Script", "Old School", 607),
        ("Script", "Varios", 608),

        ("Dingbats", "Aliens", 701),
        ("Dingbats", "Animales", 702),
        ("Dingbats", "Asia", 703),
        ("Dingbats", "Antiguo", 704),
        ("Dingbats", "Runas, Élfico", 705),
        ("Dingbats", "Esotérico", 706),
        ("Dingbats", "Fantástico", 707),
        ("Dingbats", "Terror", 708),
        ("Dingbats", "Juegos", 709),
        ("Dingbats", "Formas", 710),
        ("Dingbats", "Códigos de barras", 711),
        ("Dingbats", "Naturaleza", 712),
        ("Dingbats", "Deporte", 713),
        ("Dingbats", "Caras", 714),
        ("Dingbats", "Niñ@s", 715),
        ("Dingbats", "TV, Cine", 716),
        ("Dingbats", "Logos", 717),
        ("Dingbats", "Sexy", 718),
        ("Dingbats", "Ejército", 719),
        ("Dingbats", "Música", 720),
        ("Dingbats", "Varios", 721),

        ("Holiday", "San Valentín", 801),
        ("Holiday", "Pascua", 802),
        ("Holiday", "Halloween", 803),
        ("Holiday", "Navidad", 804),
        ("Holiday", "Varios", 805),
    ]


def get_soup(url):
    print(f"→ Cargando: {url}")
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    r.encoding = r.apparent_encoding
    return BeautifulSoup(r.text, "html.parser")


def get_total_pages(soup):
    pages = [1]

    for link in soup.find_all("a", href=True):
        href = link["href"]
        match = re.search(r"[?&]page=(\d+)", href)
        if match:
            pages.append(int(match.group(1)))

    return max(pages)


def clean_number(text):
    if not text:
        return None

    match = re.search(r"([\d\.]+)\s+descargas", text)
    if not match:
        return None

    return int(match.group(1).replace(".", ""))


def extract_fonts_from_page(soup, section, category, cat_id, page):
    rows = []

    blocks = soup.find_all("div", class_=lambda c: c and "lv1left" in c)

    for block in blocks:
        font_link = block.find("a", href=re.compile(r"\.font$"))
        if not font_link:
            continue

        strong = font_link.find("strong")
        font_name = strong.get_text(strip=True) if strong else font_link.get_text(strip=True)
        font_url = urljoin(BASE_URL, font_link["href"])

        author_name = ""
        author_url = ""

        # En el bloque lv1left, el autor suele ser el siguiente enlace después de la fuente
        links = block.find_all("a", href=True)
        for a in links:
            href = a["href"]
            if ".font" not in href and "dafont.com" in href:
                author_name = a.get_text(strip=True)
                author_url = urljoin(BASE_URL, href)
                break

        lv2 = block.find_next_sibling("div", class_=lambda c: c and "lv2right" in c)
        lv2_text = lv2.get_text(" ", strip=True) if lv2 else ""

        downloads_total = clean_number(lv2_text)

        license_text = ""
        if lv2:
            license_link = lv2.find("a", href=re.compile(r"faq\.php#copyright"))
            if license_link:
                license_text = license_link.get_text(" ", strip=True)

        preview_url = ""
        preview = block.find_next_sibling("div", class_="preview")
        if preview:
            style = preview.get("style", "")
            match = re.search(r"url\((.*?)\)", style)
            if match:
                preview_url = urljoin(BASE_URL, match.group(1).strip("'\""))

        rows.append({
            "seccion": section,
            "categoria": category,
            "cat_id": cat_id,
            "pagina": page,
            "fuente": font_name,
            "font_url": font_url,
            "autor": author_name,
            "author_url": author_url,
            "descargas_totales": downloads_total,
            "licencia": license_text,
            "preview_url": preview_url,
        })

    return rows


def save_rows(rows):
    file_exists = OUTPUT_CSV.exists()

    with open(OUTPUT_CSV, "a", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "seccion",
            "categoria",
            "cat_id",
            "pagina",
            "fuente",
            "font_url",
            "autor",
            "author_url",
            "descargas_totales",
            "licencia",
            "preview_url",
        ])

        if not file_exists:
            writer.writeheader()

        writer.writerows(rows)


def build_excel():
    df = pd.read_csv(OUTPUT_CSV)

    resumen = (
        df.groupby(["seccion", "categoria", "cat_id"], dropna=False)
        .agg(
            total_apariciones=("font_url", "count"),
            fuentes_unicas=("font_url", "nunique"),
            descargas_sumadas=("descargas_totales", "sum"),
        )
        .reset_index()
        .sort_values("total_apariciones", ascending=False)
    )

    fuentes_unicas = (
        df.groupby(["font_url", "fuente"], dropna=False)
        .agg(
            numero_categorias=("cat_id", "nunique"),
            categorias=("categoria", lambda x: "; ".join(sorted(set(map(str, x))))),
            secciones=("seccion", lambda x: "; ".join(sorted(set(map(str, x))))),
            autor=("autor", "first"),
            descargas_totales=("descargas_totales", "max"),
            licencia=("licencia", "first"),
            preview_url=("preview_url", "first"),
        )
        .reset_index()
        .sort_values("numero_categorias", ascending=False)
    )

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="Resumen_categorias", index=False)
        df.to_excel(writer, sheet_name="Fuentes_por_categoria", index=False)
        fuentes_unicas.to_excel(writer, sheet_name="Fuentes_unicas", index=False)

    wb = load_workbook(OUTPUT_XLSX)
    ws = wb["Resumen_categorias"]

    chart = BarChart()
    chart.title = "Fuentes por categoría en DaFont"
    chart.y_axis.title = "Número de fuentes"
    chart.x_axis.title = "Categoría"
    chart.height = 14
    chart.width = 30

    data = Reference(ws, min_col=4, min_row=1, max_row=min(ws.max_row, 40))
    cats = Reference(ws, min_col=2, min_row=2, max_row=min(ws.max_row, 40))

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H2")

    wb.save(OUTPUT_XLSX)

    print(f"\nExcel generado: {OUTPUT_XLSX}")


def main():
    # Para evitar mezclar pruebas anteriores
    if OUTPUT_CSV.exists():
        OUTPUT_CSV.unlink()

    categories = get_categories()

    for section, category, cat_id in categories:
        print(f"\n=== {section} > {category} ===")

        first_url = f"{THEME_URL}?cat={cat_id}"
        soup = get_soup(first_url)

        total_pages = get_total_pages(soup)
        print(f"   → {category}: {total_pages} páginas")

        total_category = 0

        for page in range(1, total_pages + 1):
            url = f"{THEME_URL}?cat={cat_id}&page={page}"
            soup = get_soup(url)

            rows = extract_fonts_from_page(soup, section, category, cat_id, page)
            save_rows(rows)

            total_category += len(rows)

            print(f"      Página {page}: {len(rows)} fuentes")

            time.sleep(1)

        print(f"TOTAL {section} > {category}: {total_category} fuentes")

    build_excel()


if __name__ == "__main__":
    main()